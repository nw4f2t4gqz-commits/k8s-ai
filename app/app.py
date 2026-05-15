import streamlit as st
from k8s_analyzer import K8sAnalyzer, GrafanaMCPClient
from rancher_client import RancherClient, MFARequired, PushPending
from kubernetes import client
import ollama
import time
import os
from datetime import datetime
import io
import shutil
import subprocess
import tempfile
from translations import CHAT_EXAMPLES, INSIGHTS_DESCRIPTIONS, get_t


def generate_fw_ticket_xls(
    namespace: str,
    policy_name: str,
    egress_vip: str,
    dest_host: str,
    dest_ip: str,
    port: str,
    justification: str,
) -> bytes:
    """Vygeneruje XLS šablonu pro FW ticket dle standardního formátu security týmu."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    # Determine VLAN from policy name: OT=54, IT=59
    policy_lower = policy_name.lower()
    if "-ot-" in policy_lower or policy_lower.startswith("ot-"):
        src_vlan = 54
        src_hostname = f"egress OT {namespace}"
    else:
        src_vlan = 59
        src_hostname = f"egress IT {namespace}"

    # Guess service type from port
    service_map = {"443": "HTTPS", "80": "HTTP", "8080": "HTTP", "8443": "HTTPS",
                   "22": "SSH", "3389": "RDP", "1433": "MSSQL", "5432": "PostgreSQL",
                   "3306": "MySQL", "5671": "AMQP", "5672": "AMQP", "4840": "OPC-UA",
                   "2400": "OPC-DA/DCOM", "21": "FTP", "25": "SMTP", "587": "SMTP"}
    service_type = service_map.get(str(port), "TCP")

    wb = Workbook()
    ws = wb.active
    ws.title = "FW Request"

    # Header style
    header_fill = PatternFill("solid", fgColor="FFC000")  # orange
    header_font = Font(bold=True, color="000000")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "No.", "Source Vlan ID", "Source Hostname", "Source IP",
        "Destination Vlan ID", "Destination Hostname", "Destination IP",
        "Protocol", "Port / Type", "Service type", "Justification of the rule"
    ]
    col_widths = [5, 14, 22, 16, 18, 36, 16, 10, 12, 14, 40]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 30

    # Data row
    row_data = [
        1, src_vlan, src_hostname, egress_vip,
        "", dest_host, dest_ip if dest_ip else dest_host,
        "TCP", port, service_type, justification
    ]
    data_fill = PatternFill("solid", fgColor="FFFFC0")  # light yellow for editable
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        if col_idx > 4:
            cell.fill = data_fill
    ws.row_dimensions[2].height = 20

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
# Ollama client - připojení na service v Kubernetes přes OLLAMA_BASE_URL
# ollama Python knihovna defaultně používá localhost:11434 - musíme explicitně předat host
OLLAMA_BASE_URL = os.environ.get('OLLAMA_BASE_URL', 'http://localhost:11434')
ollama_client = ollama.Client(host=OLLAMA_BASE_URL)

# AI Skills — system prompts loaded from ConfigMap (values.yaml → skills.*)
# Falls back to sensible defaults if not set (e.g. local dev)
_DEFAULT_INSIGHTS_SKILL = (
    "You are an expert Kubernetes L2/L3 support engineer. "
    "Answer ONLY in English. Be concise and technical. "
    "Use exact cluster data provided. Include kubectl commands where useful."
)
_DEFAULT_CHAT_SKILL = (
    "You are an expert Kubernetes L2/L3 support engineer assistant. "
    "Answer ONLY in English. "
    "When the user asks about cluster state, answer using the EXACT numbers from the LIVE CLUSTER DATA. "
    "Do NOT say 'use kubectl to find out' if the answer is already in the data. "
    "Be concise and technical."
)
AI_SKILL_INSIGHTS = os.environ.get('AI_SKILL_INSIGHTS', _DEFAULT_INSIGHTS_SKILL).strip()
AI_SKILL_CHAT = os.environ.get('AI_SKILL_CHAT', _DEFAULT_CHAT_SKILL).strip()


def has_cluster_connection() -> bool:
    """Vrací True pokud je aktivní připojení ke clusteru (Rancher session nebo kubeconfig)."""
    return bool(
        st.session_state.get('rancher_session') or
        st.session_state.get('active_kubeconfigs', [])
    )


def get_analyzer() -> K8sAnalyzer:
    """Vytvoří K8sAnalyzer pro aktuálně aktivní cluster.

    Výsledek je cachován v session_state – ApiClient se nevytváří znovu při každém rerunu.
    Cache se invaliduje při změně session (jiný cluster/login).
    """
    session = st.session_state.get('rancher_session')
    kubeconfigs = st.session_state.get('active_kubeconfigs', [])

    # Cache key – změna session = nový analyzer
    cache_key = str(session) + str(kubeconfigs[:1] if kubeconfigs else '')
    cached = st.session_state.get('_analyzer_cache')
    if cached and st.session_state.get('_analyzer_cache_key') == cache_key:
        return cached

    if session:
        analyzer = K8sAnalyzer.from_rancher(
            session['url'],
            session['cluster_id'],
            session['access_key'],
            session['secret_key'],
            session['verify_ssl'],
            bearer_token=session.get('bearer_token'),
        )
    elif kubeconfigs:
        analyzer = K8sAnalyzer(kubeconfigs[0])
    else:
        raise RuntimeError("Žádné aktivní připojení ke clusteru")

    st.session_state['_analyzer_cache'] = analyzer
    st.session_state['_analyzer_cache_key'] = cache_key
    return analyzer


st.set_page_config(page_title="K8s AI Analyzer", page_icon="🚀", layout="wide")

# Language — sidebar radio sets st.session_state['ui_lang'], defaults to 'CZ'
lang = st.session_state.get('ui_lang', 'CZ')
t = get_t(lang)

st.title("🚀 Kubernetes AI Analyzer")
st.markdown(t('app_subtitle'))

# Sidebar for configuration
with st.sidebar:
    st.radio("🌐 Jazyk / Language", ["CZ", "EN"], horizontal=True, key="ui_lang")
    st.header(t('sidebar_config'))

    # Režim připojení
    _mode_opts = [t('mode_kubeconfig'), t('mode_rancher')]
    connection_mode = st.radio(
        t('connection_mode'),
        _mode_opts,
        help=t('mode_help')
    )

    # Inicializace session state pro kubeconfigs
    if 'active_kubeconfigs' not in st.session_state:
        st.session_state.active_kubeconfigs = []
    if 'rancher_clusters' not in st.session_state:
        st.session_state.rancher_clusters = []
    if 'selected_cluster_id' not in st.session_state:
        st.session_state.selected_cluster_id = None
    if 'rancher_session' not in st.session_state:
        st.session_state.rancher_session = None

    if connection_mode == t('mode_kubeconfig'):
        kubeconfigs = st.file_uploader(
            t('upload_kubeconfig'),
            type=["yaml", "yml"],
            accept_multiple_files=True
        )

        if st.session_state.get('active_kubeconfigs', []):
            st.session_state.active_kubeconfigs = kubeconfigs
            st.success(t('kubeconfig_loaded', n=len(kubeconfigs)))

        st.session_state.rancher_clusters = []
        st.session_state.selected_cluster_id = None
        st.session_state.rancher_session = None

    else:  # Rancher Gateway
        st.subheader(t('rancher_section'))

        # --- Výběr gateway z listu ---
        _gateways_env = os.environ.get('RANCHER_GATEWAYS', '')
        _gateway_list = [u.strip() for u in _gateways_env.split(',') if u.strip()]
        _CUSTOM_LABEL = t('rancher_custom_url')
        _gateway_options = _gateway_list + [_CUSTOM_LABEL] if _gateway_list else [_CUSTOM_LABEL]

        _selected_gw = st.selectbox(
            t('rancher_gw_label'),
            _gateway_options,
            index=0,
            help=t('rancher_gw_help')
        )
        if _selected_gw == _CUSTOM_LABEL:
            rancher_url = st.text_input(
                t('rancher_url_label'),
                placeholder=t('rancher_url_ph'),
                help=t('rancher_url_help')
            )
        else:
            rancher_url = _selected_gw
            st.caption(f"🔗 {rancher_url}")

        # --- Přihlašovací metoda ---
        login_tab_up, login_tab_ping = st.tabs([t('login_tab_up'), t('login_tab_ping')])

        with login_tab_up:
            col1, col2 = st.columns(2)
            with col1:
                username = st.text_input(
                    t('username'),
                    value=os.environ.get('RANCHER_USERNAME', ''),
                    help=t('username_help'),
                    key="rancher_username",
                )
            with col2:
                password = st.text_input(
                    t('password'),
                    type="password",
                    help=t('password_help'),
                    key="rancher_password",
                )

            verify_ssl = st.checkbox(
                t('verify_ssl'),
                value=True,
                help=t('verify_ssl_help'),
                key="rancher_verify_ssl",
            )

            if st.button(t('btn_connect_rancher'), type="primary", key="btn_login_up"):
                if not rancher_url:
                    st.error(t('err_enter_url'))
                elif not username or not password:
                    st.error(t('err_enter_creds'))
                else:
                    with st.spinner(t('spinner_logging_in')):
                        try:
                            rancher = RancherClient.from_credentials(
                                rancher_url, username, password, verify_ssl
                            )
                            connected, message = rancher.test_connection()
                            if connected:
                                st.success(f"✅ {message}")
                                clusters = rancher.list_clusters()
                                st.session_state.rancher_clusters = clusters
                                st.session_state.rancher_client = rancher
                                if clusters:
                                    st.success(t('clusters_found', n=len(clusters)))
                                else:
                                    st.warning(t('no_clusters'))
                            else:
                                st.error(f"❌ {message}")
                                st.session_state.rancher_clusters = []
                        except PushPending as pp:
                            st.session_state.push_pending = pp.state
                            st.rerun()
                        except MFARequired as mfa:
                            st.session_state.mfa_pending = mfa.state
                            st.session_state.mfa_rancher_url = rancher_url
                            st.session_state.mfa_verify_ssl = verify_ssl
                            st.warning(t('mfa_required_warn'))
                            st.rerun()
                        except Exception as e:
                            st.error(t('err_generic', e=str(e)))
                            st.session_state.rancher_clusters = []

        with login_tab_ping:
            st.info(t('pingid_tab_info'))

            verify_ssl_ping = st.checkbox(
                t('verify_ssl'),
                value=True,
                key="ping_verify_ssl",
            )

            if rancher_url:
                rancher_login_url = rancher_url.rstrip('/') + "/dashboard/auth/login"
                st.link_button(
                    t('btn_open_rancher'),
                    rancher_login_url,
                    use_container_width=True,
                    type="primary",
                )
                st.caption(t('pingid_opens', url=rancher_login_url))
            else:
                st.warning(t('pingid_no_url'))

            st.markdown("---")
            st.markdown(t('pingid_after_login'))

            bearer_input = st.text_input(
                t('bearer_token_label'),
                type="password",
                placeholder="token-xxxxx:yyyyyyyyyyy",
                help=t('bearer_token_help'),
                key="ping_bearer_input",
            )
            if st.button(t('btn_login_token'), type="primary", key="btn_ping_token_login"):
                if not rancher_url:
                    st.error(t('err_enter_url'))
                elif not bearer_input:
                    st.error(t('err_enter_token'))
                else:
                    with st.spinner(t('spinner_verify_token')):
                        try:
                            rancher = RancherClient(
                                rancher_url=rancher_url,
                                verify_ssl=verify_ssl_ping,
                                bearer_token=bearer_input.strip(),
                            )
                            connected, message = rancher.test_connection()
                            if connected:
                                st.success(f"✅ {message}")
                                clusters = rancher.list_clusters()
                                st.session_state.rancher_clusters = clusters
                                st.session_state.rancher_client = rancher
                                if clusters:
                                    st.success(t('clusters_found', n=len(clusters)))
                                else:
                                    st.warning(t('no_clusters'))
                                st.rerun()
                            else:
                                st.error(f"❌ {message}")
                        except Exception as e:
                            st.error(f"❌ {e}")

        verify_ssl = st.session_state.get("rancher_verify_ssl", True)

        # PingID push pending — čeká na schválení na mobilu
        if st.session_state.get('push_pending'):
            _pp = st.session_state.push_pending
            _ppm_ok    = bool(_pp.get('ppm_params'))
            _ppm_err   = _pp.get('ppm_error')
            if _ppm_ok:
                st.success(t('push_sent_ok'))
                st.info(t('push_sent_info'))
            else:
                st.error(t('push_failed_err'))
                if _ppm_err:
                    st.caption(t('push_failed_caption', e=_ppm_err[:200]))
            col_p1, col_p2, col_p3 = st.columns([3, 2, 1])
            with col_p1:
                if st.button(t('btn_push_approved'), type="primary", key="btn_push_approved"):
                    with st.spinner(t('spinner_verifying')):
                        try:
                            token = RancherClient._saml_check_push(st.session_state.push_pending)
                            _push_rancher_url = st.session_state.push_pending.get('rancher_url', rancher_url)
                            _push_verify_ssl  = st.session_state.push_pending.get('verify_ssl', True)
                            rancher = RancherClient(
                                rancher_url=_push_rancher_url,
                                verify_ssl=_push_verify_ssl,
                                bearer_token=token,
                            )
                            connected, message = rancher.test_connection()
                            if connected:
                                st.session_state.push_pending = None
                                st.success(f"✅ {message}")
                                clusters = rancher.list_clusters()
                                st.session_state.rancher_clusters = clusters
                                st.session_state.rancher_client = rancher
                                if clusters:
                                    st.success(t('clusters_found', n=len(clusters)))
                                else:
                                    st.warning(t('no_clusters'))
                                st.rerun()
                            else:
                                st.error(f"❌ {message}")
                        except MFARequired as mfa:
                            st.session_state.push_pending = None
                            st.session_state.mfa_pending = mfa.state
                            st.session_state.mfa_rancher_url = rancher_url
                            st.session_state.mfa_verify_ssl = verify_ssl
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ {e}")
            with col_p2:
                if st.button(t('btn_push_retry'), key="btn_push_retry"):
                    st.session_state.push_pending = None
                    st.info(t('push_retry_info'))
                    st.rerun()
            with col_p3:
                if st.button(t('btn_cancel'), key="btn_push_cancel"):
                    st.session_state.push_pending = None
                    st.rerun()

            # Debug expander — zobrazit HTML stránky pro diagnostiku
            _page_html = st.session_state.push_pending.get('page_html', '') if st.session_state.get('push_pending') else ''
            if _page_html:
                with st.expander(t('push_debug_expander')):
                    st.caption(t('push_debug_pending_url', url=st.session_state.push_pending.get('pending_url', '?')))
                    import re as _re
                    _scripts = _re.findall(r'<script[^>]*>(.*?)</script>', _page_html, _re.DOTALL | _re.IGNORECASE)
                    _urls_found = []
                    for _s in _scripts:
                        _urls_found += _re.findall(r"""['"](/idp/[^'"?#\s]{5,})['""]""", _s)
                    if _urls_found:
                        st.markdown(t('push_debug_js_urls'))
                        for _u in list(dict.fromkeys(_urls_found))[:15]:
                            st.code(_u)
                    st.markdown(t('push_debug_html'))
                    st.code(_page_html[:3000], language="html")

        # MFA OTP challenge (PingID)
        if st.session_state.get('mfa_pending'):
            st.warning(t('mfa_warning'))
            otp_code = st.text_input(
                t('otp_label'), key="mfa_otp_input",
                placeholder="123456", max_chars=8,
                help=t('otp_help')
            )
            col_otp1, col_otp2 = st.columns([1, 1])
            with col_otp1:
                if st.button(t('btn_verify_otp'), type="primary", key="mfa_submit"):
                    if not otp_code:
                        st.error(t('err_enter_otp'))
                    else:
                        with st.spinner(t('spinner_verify_otp')):
                            try:
                                token = RancherClient._saml_login_complete(
                                    st.session_state.mfa_pending, otp_code
                                )
                                rancher = RancherClient(
                                    rancher_url=st.session_state.mfa_rancher_url,
                                    verify_ssl=st.session_state.mfa_verify_ssl,
                                    bearer_token=token,
                                )
                                connected, message = rancher.test_connection()
                                if connected:
                                    st.session_state.mfa_pending = None
                                    st.success(f"✅ {message}")
                                    clusters = rancher.list_clusters()
                                    st.session_state.rancher_clusters = clusters
                                    st.session_state.rancher_client = rancher
                                    if clusters:
                                        st.success(t('clusters_found', n=len(clusters)))
                                    else:
                                        st.warning(t('no_clusters'))
                                    st.rerun()
                                else:
                                    st.error(f"❌ {message}")
                            except Exception as e:
                                st.error(f"❌ MFA chyba: {e}")
            with col_otp2:
                if st.button(t('btn_cancel'), key="mfa_cancel"):
                    st.session_state.mfa_pending = None
                    st.rerun()

        # Výběr clusteru
        if st.session_state.rancher_clusters:
            st.markdown("---")
            st.subheader(t('available_clusters'))

            cluster_options = [
                f"{cluster['name']} ({cluster['state']}) - {cluster['version']}"
                for cluster in st.session_state.rancher_clusters
            ]

            selected_cluster = st.selectbox(
                t('select_cluster'),
                cluster_options,
                help=t('select_cluster_help')
            )

            if selected_cluster:
                # Najít vybraný cluster
                cluster_name = selected_cluster.split(" (")[0]
                for cluster in st.session_state.rancher_clusters:
                    if cluster['name'] == cluster_name:
                        st.session_state.selected_cluster_id = cluster['id']

                        # Zobrazit info o clusteru
                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric(t('metric_state'), cluster['state'])
                        with col2:
                            st.metric(t('metric_version'), cluster['version'])

                        # Načíst kubeconfig pro vybraný cluster
                        if st.button(t('btn_load_kubeconfig')):
                            # Vždy vymazat staré kubeconfig data před novým načtením
                            st.session_state.active_kubeconfigs = []
                            with st.spinner(t('spinner_kubeconfig')):
                                rancher = st.session_state.rancher_client
                                kubeconfig_str = None
                                kubeconfig_method = ""

                                # 1. Pokus: Rancher generateKubeconfig (preferovaná metoda)
                                try:
                                    kubeconfig_str = rancher.get_kubeconfig(cluster['id'])
                                    # Zkontrolovat, zda kubeconfig nepoužívá exec credentials
                                    # (kubectl plugin) - Python klient to neumob načíst bez CLI
                                    if rancher.kubeconfig_has_exec_credentials(kubeconfig_str):
                                        st.warning(t('warn_exec_creds'))
                                        kubeconfig_str = None
                                    else:
                                        kubeconfig_method = "Rancher generateKubeconfig"
                                except Exception as e:
                                    err = str(e)
                                    st.warning(t('warn_gen_failed', e=err))
                                    if '403' in err or 'Forbidden' in err:
                                        st.info(
                                            t('info_403_key')
                                        )

                                # 2. Záložní metoda: proxy kubeconfig s Basic Auth
                                # (username/password = access_key:secret_key)
                                # Rancher proxy /k8s/clusters/<id> akceptuje Basic Auth
                                if kubeconfig_str is None:
                                    try:
                                        kubeconfig_str = rancher.generate_proxy_kubeconfig(
                                            cluster['id'], cluster['name']
                                        )
                                        kubeconfig_method = "Rancher Proxy (Basic Auth)"
                                    except Exception as e:
                                        st.error(t('err_proxy_kube', e=str(e)))

                                if kubeconfig_str:
                                    kubeconfig_bytes = io.BytesIO(kubeconfig_str.encode())
                                    kubeconfig_bytes.name = f"{cluster['name']}.yaml"
                                    st.session_state.active_kubeconfigs = [kubeconfig_bytes]
                                    # Uložit Rancher session – K8sAnalyzer.from_rancher() použije
                                    # přímé Basic Auth připojení místo kubeconfig souboru
                                    st.session_state.rancher_session = {
                                        'url': rancher.rancher_url,
                                        'cluster_id': cluster['id'],
                                        'cluster_name': cluster['name'],
                                        'access_key': rancher.access_key,
                                        'secret_key': rancher.secret_key,
                                        'bearer_token': rancher.bearer_token,
                                        'verify_ssl': rancher.verify_ssl,
                                    }
                                    st.success(t('kubeconfig_ok', m=kubeconfig_method))

                                    # Debug: zobrazit část kubeconfig pro kontrolu
                                    with st.expander(t('kubeconfig_preview')):
                                        preview = kubeconfig_str[:600] + "..." if len(kubeconfig_str) > 600 else kubeconfig_str
                                        st.code(preview, language="yaml")
                                else:
                                    st.error(t('kubeconfig_fail'))
                                    st.session_state.active_kubeconfigs = []
                                    st.session_state.rancher_session = None
                        break

        # Vyčistit klasické kubeconfigs
        kubeconfigs = st.session_state.active_kubeconfigs

    st.markdown("---")

    # AI Model selection — default: Qwen2.5:1.5B (fastest in benchmark 2026-03-17)
    model_options = ["Qwen2.5:1.5B", "llama3.2:1b", "phi3.5:latest", "qwen3.5:2b", "qwen3:8b", "mistral:7b"]
    selected_model = st.selectbox(t('ai_model'), model_options)

    # Check if Ollama is running
    try:
        ollama_client.list()
        st.success(t('ollama_ok', url=OLLAMA_BASE_URL))
    except Exception as e:
        st.error(t('ollama_err', url=OLLAMA_BASE_URL, e=e))

# Main content
tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([t('tab_overview'), t('tab_pods'), t('tab_events'), t('tab_insights'), t('tab_chat'), t('tab_k8sgpt'), t('tab_logs'), t('tab_egress')])

with tab1:
    st.header(t('h_cluster_overview'))

    if has_cluster_connection():
        try:
            analyzer = get_analyzer()
        except Exception as e:
            st.error(t('err_init', e=str(e)))
            st.stop()

        # Test connection
        connected, message = analyzer.test_connection()
        if connected:
            st.success(message)
        else:
            st.error(message)
            st.stop()  # Stop execution if connection fails

        col1, col2, col3 = st.columns(3)

        with col1:
            nodes = analyzer.get_nodes()
            st.metric(t('metric_nodes'), len(nodes))

        with col2:
            pods = analyzer.get_pods()
            st.metric(t('metric_total_pods'), len(pods))

        with col3:
            namespaces = analyzer.get_namespaces()
            st.metric(t('metric_namespaces'), len(namespaces))

        # Node status
        st.subheader(t('h_node_status'))
        node_data = []
        for node in nodes:
            node_data.append({
                'Name': node.metadata.name,
                'Status': node.status.conditions[-1].type if node.status.conditions else 'Unknown',
                'CPU': node.status.capacity.get('cpu', 'N/A'),
                'Memory': node.status.capacity.get('memory', 'N/A')
            })
        st.table(node_data)

    else:
        st.info(t('connect_prompt'))

with tab2:
    st.header(t('h_pod_analysis'))

    if has_cluster_connection():
        analyzer = get_analyzer()

        # Pod status overview
        pods = analyzer.get_pods()
        status_counts = {}
        for pod in pods:
            status = pod.status.phase
            status_counts[status] = status_counts.get(status, 0) + 1

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric(t('pod_running'), status_counts.get('Running', 0))
        with col2:
            st.metric(t('pod_pending'), status_counts.get('Pending', 0))
        with col3:
            st.metric(t('pod_failed'), status_counts.get('Failed', 0))
        with col4:
            st.metric(t('pod_succeeded'), status_counts.get('Succeeded', 0))

        # Problematic pods
        problematic_pods = [pod for pod in pods if pod.status.phase in ['Pending', 'Failed', 'CrashLoopBackOff']]
        if problematic_pods:
            st.subheader(t('h_problematic_pods'))
            prob_data = []
            for pod in problematic_pods:
                restarts = 0
                if pod.status.container_statuses:
                    restarts = pod.status.container_statuses[0].restart_count
                prob_data.append({
                    'Name': pod.metadata.name,
                    'Namespace': pod.metadata.namespace,
                    'Status': pod.status.phase,
                    'Restarts': restarts
                })
            st.table(prob_data)

with tab3:
    st.header(t('h_events'))

    if has_cluster_connection():
        analyzer = get_analyzer()

        # Recent events
        st.subheader(t('h_recent_events'))
        events = analyzer.get_recent_events(hours=1)

        if events:
            event_data = []
            for event in events[:20]:  # Show last 20 events
                event_data.append({
                    'Time': event['time'].strftime('%H:%M:%S') if hasattr(event['time'], 'strftime') else str(event['time']),
                    'Type': event['type'],
                    'Reason': event['reason'],
                    'Message': event['message'][:100] + '...' if len(event['message']) > 100 else event['message'],
                    'Source': event['source'],
                    'Object': event['object']
                })
            st.table(event_data)
        else:
            st.info(t('no_recent_events'))

        # Resource utilization
        st.subheader(t('h_resource_util'))
        nodes = analyzer.get_nodes()
        pods = analyzer.get_pods()

        total_cpu = 0
        total_memory = 0
        used_cpu = 0
        used_memory = 0

        for node in nodes:
            if node.status.capacity:
                cpu = node.status.capacity.get('cpu', '0')
                memory = node.status.capacity.get('memory', '0')

                # Parse CPU (can be in cores or millicores)
                if cpu.endswith('m'):
                    total_cpu += float(cpu[:-1]) / 1000
                else:
                    total_cpu += float(cpu) if cpu.isdigit() else 0

                # Parse memory (convert to Mi)
                if memory.endswith('Ki'):
                    total_memory += float(memory[:-2]) / 1024
                elif memory.endswith('Mi'):
                    total_memory += float(memory[:-2])
                elif memory.endswith('Gi'):
                    total_memory += float(memory[:-2]) * 1024

        # Calculate used resources from pods
        for pod in pods:
            if pod.spec.containers and pod.status.phase == 'Running':
                for container in pod.spec.containers:
                    if container.resources.requests:
                        cpu_req = container.resources.requests.get('cpu', '0')
                        mem_req = container.resources.requests.get('memory', '0')

                        if cpu_req.endswith('m'):
                            used_cpu += float(cpu_req[:-1]) / 1000
                        else:
                            used_cpu += float(cpu_req) if cpu_req.isdigit() else 0

                        if mem_req.endswith('Mi'):
                            used_memory += float(mem_req[:-2])
                        elif mem_req.endswith('Gi'):
                            used_memory += float(mem_req[:-2]) * 1024

        col1, col2 = st.columns(2)
        with col1:
            cpu_usage = (used_cpu / total_cpu * 100) if total_cpu > 0 else 0
            st.metric(t('cpu_usage'), f"{cpu_usage:.1f}%", f"{used_cpu:.2f}/{total_cpu:.2f} cores")
            st.progress(min(cpu_usage / 100, 1.0))

        with col2:
            mem_usage = (used_memory / total_memory * 100) if total_memory > 0 else 0
            st.metric(t('mem_usage'), f"{mem_usage:.1f}%", f"{used_memory:.0f}/{total_memory:.0f} Mi")
            st.progress(min(mem_usage / 100, 1.0))

with tab4:
    @st.fragment
    def render_ai_insights():
        st.header(t('h_ai_insights'))

        if not has_cluster_connection():
            st.info(t('not_connected'))
            return

        analyzer = get_analyzer()

        analysis_type = st.selectbox(
            t('analysis_type'),
            ["General Cluster Health", "Problematic Pods Analysis", "Resource Optimization", "Security Check"],
            key="ai_analysis_type",
        )

        # --- Popis vybraneho typu analyzy ---
        _insights_desc = INSIGHTS_DESCRIPTIONS[lang]
        st.caption(_insights_desc.get(analysis_type, ""))
        st.divider()

        if st.button(t('btn_generate'), key="ai_generate_btn"):
            # 1. Sbírání K8s dat
            with st.spinner(t('spinner_loading_data')):
                cluster_info = analyzer.get_cluster_summary()
                events = analyzer.get_recent_events(hours=2)
                problematic_pods = [pod for pod in analyzer.get_pods() if pod.status.phase in ['Pending', 'Failed', 'CrashLoopBackOff']]

            # 2. Sestavení promptu
            if analysis_type == "General Cluster Health":
                prompt = (
                    f"Analyze Kubernetes cluster health for L2 support.\n"
                    f"Nodes: {cluster_info['nodes']}, Pods: {cluster_info['pods']}, "
                    f"Namespaces: {cluster_info['namespaces']}, Problematic pods: {cluster_info['problematic_pods']}, "
                    f"Recent events: {len(events)} in last 2h.\n"
                    "Provide: 1) Health assessment 2) Critical issues 3) Recommendations. Be concise."
                )
            elif analysis_type == "Problematic Pods Analysis":
                pod_list = "\n".join([f"- {p.metadata.name} ({p.metadata.namespace}): {p.status.phase}" for p in problematic_pods[:5]])
                prompt = (
                    f"Analyze {len(problematic_pods)} problematic Kubernetes pods:\n{pod_list}\n"
                    f"Pod-related events: {len([e for e in events if 'pod' in e['object'].lower()])}\n"
                    "Provide: 1) Root cause 2) Troubleshooting steps 3) kubectl commands 4) Fixes. Be concise."
                )
            elif analysis_type == "Resource Optimization":
                prompt = (
                    f"Analyze Kubernetes resource utilization.\n"
                    f"Nodes: {cluster_info['nodes']}, Pods: {cluster_info['pods']}, "
                    f"Distribution: {cluster_info['pod_status_distribution']}\n"
                    "Provide: 1) Utilization assessment 2) Over/under-provisioned 3) Optimization 4) Scaling. Be concise."
                )
            else:
                prompt = (
                    f"Security analysis of Kubernetes cluster.\n"
                    f"Namespaces: {cluster_info['namespaces']}, Pods: {cluster_info['pods']}, Events: {len(events)}\n"
                    "Provide: 1) Vulnerabilities 2) RBAC issues 3) Network security 4) Best practices. Be concise."
                )

            # 3. Streaming — BEZ st.spinner()
            is_thinking_model = any(x in selected_model.lower() for x in ['qwen3', 'qwq', 'deepseek-r'])
            if is_thinking_model:
                prompt = "/no_think\n" + prompt

            system_prompt = AI_SKILL_INSIGHTS

            status_placeholder = st.empty()
            status_placeholder.caption(t('model_starting', m=selected_model))
            result_placeholder = st.empty()
            full_response = ""
            think_buf = ""
            in_think = False
            token_count = 0
            try:
                stream = ollama_client.generate(
                    model=selected_model,
                    system=system_prompt,
                    prompt=prompt,
                    stream=True,
                    options={
                        "num_predict": 500,
                        "temperature": 0.1,
                        "top_p": 0.9,
                        "stop": ["\n\n\n", "Human:", "User:", "Assistant:", "Translate", "Rédigez", "指令"],
                    },
                )
                for chunk in stream:
                    token = chunk.response if hasattr(chunk, 'response') else chunk.get('response', '')
                    if not token:
                        continue
                    token_count += 1
                    if '<think>' in token:
                        in_think = True
                    if in_think:
                        think_buf += token
                        status_placeholder.caption(t('model_thinking', n=token_count))
                        if '</think>' in token:
                            in_think = False
                            status_placeholder.caption(t('thinking_done'))
                        continue
                    full_response += token
                    status_placeholder.empty()
                    result_placeholder.markdown(full_response + " ▮")
                result_placeholder.markdown(full_response)
                st.caption(t('analysis_done', m=selected_model, n=len(full_response.split())))
            except Exception as e:
                st.error(t('analysis_error', e=str(e)))

    render_ai_insights()

with tab5:
    @st.fragment
    def render_chat_tab():
        st.header(t('h_ai_chat'))

        if "messages" not in st.session_state:
            st.session_state.messages = []

        # --- Load cluster context ONCE per cluster, invalidate on cluster change ---
        _current_cache_key = st.session_state.get('_analyzer_cache_key', '')

        # Determine if Grafana Prometheus data is relevant for the currently selected cluster.
        # GRAFANA_CLUSTER env var contains the name of the cluster monitored by this Grafana.
        # If a different cluster is selected (Rancher mode), Prometheus data would be wrong.
        _selected_id = st.session_state.get('selected_cluster_id')
        _rancher_clusters = st.session_state.get('rancher_clusters', [])
        _current_cluster_name = next(
            (c['name'] for c in _rancher_clusters if c['id'] == _selected_id), None
        )

        if (st.session_state.get('chat_cluster_context_key') != _current_cache_key
                or "chat_cluster_context" not in st.session_state):
            if has_cluster_connection():
                with st.spinner(t('spinner_ctx_loading')):
                    try:
                        _analyzer = get_analyzer()
                        _ci = _analyzer.get_cluster_summary()
                        _events = _analyzer.get_recent_events(hours=1)
                        _nodes = _analyzer.get_nodes()
                        _dist = _ci['pod_status_distribution']
                        _running = _dist.get('Running', 0)
                        _bad = [p for p in _analyzer.get_pods()
                                if p.status.phase in ['Pending', 'Failed', 'CrashLoopBackOff']]
                        _bad_lines = "\n".join(
                            [f"  - {p.metadata.name} ({p.metadata.namespace}): {p.status.phase}"
                             for p in _bad[:10]]
                        ) if _bad else "  (none)"
                        _dist_lines = "\n".join(
                            [f"  - {phase}: {count}" for phase, count in sorted(_dist.items())]
                        )
                        # Node capacity — allocatable.pods per node
                        _node_lines = []
                        _total_max_pods = 0
                        for _n in _nodes:
                            _alloc = getattr(_n.status, 'allocatable', {}) or {}
                            _cap = getattr(_n.status, 'capacity', {}) or {}
                            _max = int(_alloc.get('pods', _cap.get('pods', 0)) or 0)
                            _total_max_pods += _max
                            _ready = any(
                                c.type == 'Ready' and c.status == 'True'
                                for c in (_n.status.conditions or [])
                            )
                            _node_lines.append(
                                f"  - {_n.metadata.name}: max {_max} pods, Ready={_ready}"
                            )
                        _node_detail = "\n".join(_node_lines) if _node_lines else "  (unknown)"
                        st.session_state.chat_cluster_context = (
                            f"LIVE CLUSTER DATA (fetched {datetime.now().strftime('%H:%M:%S')}):\n"
                            f"- Nodes: {_ci['nodes']}\n"
                            f"- Total pods: {_ci['pods']}\n"
                            f"- Running pods: {_running}\n"
                            f"- Problematic pods: {_ci['problematic_pods']}\n"
                            f"- Namespaces: {_ci['namespaces']}\n"
                            f"- Events last 1h: {len(_events)}\n"
                            f"- Max pods capacity (sum of all nodes allocatable): {_total_max_pods}\n"
                            f"- Pod status breakdown:\n{_dist_lines}\n"
                            f"- Node details:\n{_node_detail}\n"
                            f"- Problematic pod list:\n{_bad_lines}"
                        )
                        st.session_state.chat_cluster_context_time = datetime.now()
                        st.session_state.chat_cluster_context_key = _current_cache_key

                        # --- Grafana MCP: real-time Prometheus metrics + alerts ---
                        # Inject cluster label into PromQL when Prometheus is multi-cluster
                        # (same pattern as Loki cluster label filtering)
                        _mcp = GrafanaMCPClient.detect()
                        if _mcp:
                            _prom_uid = _mcp.get_prometheus_datasource_uid()
                            _mcp_extras = []
                            if _prom_uid:
                                # Detect multi-cluster Prometheus and pass cluster filter
                                _prom_has_cluster = _mcp.has_cluster_label(_prom_uid)
                                _query_cluster = _current_cluster_name if (_prom_has_cluster and _current_cluster_name) else None
                                _gmetrics = _mcp.get_cluster_metrics_summary(_prom_uid, cluster=_query_cluster)
                                if _gmetrics.get("cpu_usage_per_node"):
                                    _mcp_extras.append(f"Node CPU usage (%):\n{_gmetrics['cpu_usage_per_node']}")
                                if _gmetrics.get("memory_usage_per_node"):
                                    _mcp_extras.append(f"Node Memory usage (%):\n{_gmetrics['memory_usage_per_node']}")
                                if _gmetrics.get("top_cpu_pods"):
                                    _mcp_extras.append(f"Top 5 pods by CPU (millicores):\n{_gmetrics['top_cpu_pods']}")
                                if _gmetrics.get("top_memory_pods"):
                                    _mcp_extras.append(f"Top 5 pods by Memory (MiB):\n{_gmetrics['top_memory_pods']}")
                            _alerts = _mcp.get_firing_alerts()
                            if _alerts:
                                _mcp_extras.append(f"Grafana Alerts:\n{_alerts[:600]}")
                            if _mcp_extras:
                                _cluster_tag = f" ({_current_cluster_name})" if _current_cluster_name else ""
                                st.session_state.chat_cluster_context += (
                                    f"\n\nLIVE PROMETHEUS METRICS (via Grafana{_cluster_tag}):\n"
                                    + "\n\n".join(_mcp_extras)
                                )
                                st.session_state.chat_cluster_context_time = datetime.now()
                    except Exception as _e:
                        st.session_state.chat_cluster_context = t('ctx_unavailable')
                        st.session_state.chat_cluster_context_time = datetime.now()
                        st.session_state.chat_cluster_context_key = _current_cache_key
            else:
                st.session_state.chat_cluster_context = t('ctx_not_connected')
                st.session_state.chat_cluster_context_time = datetime.now()
                st.session_state.chat_cluster_context_key = _current_cache_key

        # --- Status bar + Refresh button ---
        _col_info, _col_ref = st.columns([5, 1])
        with _col_info:
            if not GrafanaMCPClient.detect():
                _mcp_status = t('grafana_offline')
            elif _current_cluster_name:
                _mcp_status = t('prometheus_ok_cluster', c=_current_cluster_name)
            else:
                _mcp_status = t('prometheus_ok')
            _ctx_time = st.session_state.get('chat_cluster_context_time')
            _time_str = _ctx_time.strftime('%H:%M:%S') if hasattr(_ctx_time, 'strftime') else '?'
            st.caption(t('ctx_status', t=_time_str, s=_mcp_status))
        with _col_ref:
            if st.button("🔄", key="chat_refresh_btn", help=t('btn_refresh_ctx')):
                st.session_state.pop("chat_cluster_context", None)
                st.session_state.pop("chat_cluster_context_time", None)
                st.session_state.pop("chat_cluster_context_key", None)
                st.rerun(scope="fragment")

        # --- Příklady dotazů ---
        with st.expander(t('expander_examples'), expanded=not st.session_state.get('messages')):
            _examples = CHAT_EXAMPLES[lang]
            _ex_cols = st.columns(2)
            for _i, (_label, _question) in enumerate(_examples):
                if _ex_cols[_i % 2].button(_label, key=f"chat_ex_{_i}", help=_question, use_container_width=True):
                    st.session_state['_chat_pending'] = _question
                    st.rerun(scope="fragment")

        # Pick up pending suggestion before rendering text input
        _auto_question = st.session_state.pop('_chat_pending', None)

        chat_container = st.container()
        with chat_container:
            for message in st.session_state.messages:
                with st.chat_message(message["role"]):
                    st.markdown(message["content"])

        # Chat input
        col1, col2 = st.columns([4, 1])
        with col1:
            user_prompt = st.text_input(t('chat_placeholder'), key="chat_input", label_visibility="collapsed")
        with col2:
            send_button = st.button(t('btn_send'), use_container_width=True, key="chat_send_btn")

        effective_prompt = _auto_question or (user_prompt if send_button else None)
        if effective_prompt:
            user_prompt = effective_prompt
            st.session_state.messages.append({"role": "user", "content": user_prompt})
            with chat_container:
                with st.chat_message("user"):
                    st.markdown(user_prompt)

            with chat_container:
                with st.chat_message("assistant"):
                    chat_status = st.empty()
                    response_placeholder = st.empty()
                    full_response = ""
                    think_buf = ""
                    in_think = False
                    token_count = 0
                    try:
                        cluster_context = st.session_state.get("chat_cluster_context", "")
                        is_thinking_model = any(x in selected_model.lower() for x in ['qwen3', 'qwq', 'deepseek-r'])
                        no_think_prefix = "/no_think\n" if is_thinking_model else ""
                        full_prompt = (
                            no_think_prefix +
                            f"{cluster_context}\n\n"
                            "IMPORTANT: Use the exact numbers from the LIVE CLUSTER DATA above to answer factual questions. "
                            "Do NOT say 'use kubectl to find out' if the answer is already in the data above.\n\n"
                            f"Question: {user_prompt}"
                        )
                        chat_system_prompt = AI_SKILL_CHAT
                        stream = ollama_client.generate(
                            model=selected_model,
                            system=chat_system_prompt,
                            prompt=full_prompt,
                            stream=True,
                            options={
                                "num_predict": 600,
                                "temperature": 0.1,
                                "top_p": 0.9,
                                "stop": ["\n\n\n", "Human:", "User:", "Assistant:", "Question:", "<|", "Translate", "Rédigez", "指令"],
                            },
                        )
                        for chunk in stream:
                            token = chunk.response if hasattr(chunk, 'response') else chunk.get('response', '')
                            if not token:
                                continue
                            token_count += 1
                            if '<think>' in token:
                                in_think = True
                            if in_think:
                                think_buf += token
                                chat_status.caption(t('model_thinking', n=token_count))
                                if '</think>' in token:
                                    in_think = False
                                    chat_status.caption(t('thinking_done'))
                                continue
                            full_response += token
                            chat_status.empty()
                            response_placeholder.markdown(full_response + " ▮")
                        response_placeholder.markdown(full_response)
                        st.session_state.messages.append({"role": "assistant", "content": full_response})
                    except Exception as e:
                        error_msg = t('chat_error', e=str(e))
                        st.error(error_msg)
                        st.session_state.messages.append({"role": "assistant", "content": error_msg})

            st.rerun(scope="fragment")

        if st.button(t('btn_clear_chat'), key="chat_clear_btn"):
            st.session_state.messages = []
            st.rerun(scope="fragment")

    render_chat_tab()

with tab6:
    st.header(t('h_k8sgpt'))

    st.markdown(t('k8sgpt_desc'))

    if not has_cluster_connection():
        st.info(t('connect_prompt'))
    else:
        analyzer = get_analyzer()

        # Load Results via kubernetes custom objects API
        col_refresh, col_ns, col_filter = st.columns([1, 2, 2])
        with col_ns:
            result_ns = st.selectbox(
                "Namespace",
                ["all"] + [ns.metadata.name for ns in analyzer.get_namespaces()],
                index=0,
                key="k8sgpt_ns"
            )
        with col_filter:
            kind_filter = st.multiselect(
                t('filter_by_kind'),
                ["Pod", "Service", "Deployment", "StatefulSet", "DaemonSet", "ConfigMap", "Secret", "PersistentVolumeClaim", "Ingress"],
                default=[],
                key="k8sgpt_kind"
            )
        with col_refresh:
            st.write("")  # spacing
            do_load = st.button(t('btn_load_results'), type="primary")

        if do_load:
            with st.spinner(t('spinner_k8sgpt')):
                try:
                    custom_api = client.CustomObjectsApi(analyzer.v1.api_client)
                    if result_ns == "all":
                        resp = custom_api.list_cluster_custom_object(
                            group="core.k8sgpt.ai",
                            version="v1alpha1",
                            plural="results",
                        )
                    else:
                        resp = custom_api.list_namespaced_custom_object(
                            group="core.k8sgpt.ai",
                            version="v1alpha1",
                            plural="results",
                            namespace=result_ns,
                        )

                    items = resp.get("items", [])

                    # Filtrovat dle kind
                    if kind_filter:
                        items = [r for r in items if r.get("spec", {}).get("kind", "") in kind_filter]

                    if not items:
                        st.success(t('k8sgpt_no_issues'))
                        st.info(t('k8sgpt_hint'))
                    else:
                        st.warning(t('k8sgpt_issues_found', n=len(items)))

                        for item in items:
                            spec = item.get("spec", {})
                            meta = item.get("metadata", {})
                            kind = spec.get("kind", "Unknown")
                            name = spec.get("name", meta.get("name", "?"))
                            parent = spec.get("parentObject", "")
                            details = spec.get("details", "")
                            errors = spec.get("error", [])
                            ns_label = meta.get("namespace", "cluster-wide")

                            # Ikona dle severity
                            icon = "🔴" if errors else "🟡"
                            with st.expander(f"{icon} **{kind}** `{name}` — {ns_label}"):
                                if parent:
                                    st.caption(f"Parent: {parent}")
                                if details:
                                    st.markdown(t('k8sgpt_ai_analysis') + f"\n\n{details}")
                                if errors:
                                    st.markdown(t('k8sgpt_errors'))
                                    for e in errors:
                                        st.code(e.get("text", ""), language="")

                except client.exceptions.ApiException as e:
                    if e.status == 403:
                        st.error(
        t('k8sgpt_403')
                        )
                    elif e.status == 404:
                        st.error(
        t('k8sgpt_no_crd')
                        )
                    else:
                        st.error(t('k8sgpt_api_err', e=e))
                except Exception as e:
                    st.error(t('err_generic', e=str(e)))

        # Status K8sGPT CR
        with st.expander(t('k8sgpt_cr_status')):
            if not has_cluster_connection():
                st.info(t('k8sgpt_not_connected'))
            else:
                if st.button(t('btn_load_cr'), key="load_k8sgpt_cr"):
                    try:
                        analyzer2 = get_analyzer()
                        custom_api2 = client.CustomObjectsApi(analyzer2.v1.api_client)
                        # K8sGPT CR je namespace-scoped — hledáme ve všech NS
                        k8sgpts = custom_api2.list_cluster_custom_object(
                            group="core.k8sgpt.ai",
                            version="v1alpha1",
                            plural="k8sgpts",
                        )
                        if not k8sgpts.get("items"):
                            st.warning(t('k8sgpt_cr_none'))
                        else:
                            for k in k8sgpts["items"]:
                                meta = k.get("metadata", {})
                                st_obj = k.get("status", {})
                                st.json({"name": meta.get("name"), "namespace": meta.get("namespace"), "status": st_obj})
                    except client.exceptions.ApiException as e:
                        if e.status == 404:
                            st.warning(
    t('k8sgpt_cr_no_crd')
                            )
                        elif e.status == 403:
                            st.error(t('k8sgpt_cr_403'))
                        else:
                            st.error(t('k8sgpt_cr_api_err', s=e.status, r=e.reason))
                    except Exception as e:
                        st.warning(t('k8sgpt_cr_warn', e=e))

@st.fragment
def render_logs_tab():
    st.header(t('h_logs'))

    if not has_cluster_connection():
        st.info(t('not_connected'))
        return

    analyzer = get_analyzer()

    # cluster_name pro Loki label (napr. czplskbe1001, eudrpkbe0001)
    _session = st.session_state.get('rancher_session', {})
    loki_cluster = _session.get('cluster_name', '') if _session else ''

    # --- Loki detekce (jednou za session, ne při každém rerun) ---
    if 'log_loki_source' not in st.session_state:
        loki = analyzer.get_loki_client(verify_ssl=_session.get('verify_ssl', True))
        if loki and loki.is_available():
            st.session_state['log_loki_client'] = loki
            st.session_state['log_loki_source'] = True
            # Namespacy filtrové podle clusteru (pokud Loki má cluster label)
            ns_vals = loki.get_label_values('namespace', cluster=loki_cluster)
            st.session_state['log_loki_namespaces'] = set(ns_vals)
        else:
            st.session_state['log_loki_client'] = None
            st.session_state['log_loki_source'] = False
            st.session_state['log_loki_namespaces'] = set()

    loki = st.session_state.get('log_loki_client')
    loki_namespaces = st.session_state.get('log_loki_namespaces', set())
    log_source = 'loki' if st.session_state['log_loki_source'] else 'k8s'

    if log_source == 'loki':
        cluster_label_info = f" | cluster: `{loki_cluster}`" if loki_cluster and getattr(loki, 'has_cluster_label', False) else ""
        st.success(t('loki_ok', url=loki.url, cinfo=cluster_label_info, n=len(loki_namespaces)))
        with st.expander(t('loki_diag')):
            if st.button(t('loki_show_labels')):
                labels = loki.get_labels()
                st.write(t('loki_labels_title'), labels)
            if st.button(t('loki_show_ns')):
                st.write(t('loki_ns_title'), sorted(loki_namespaces))
            if 'log_ns' in st.session_state:
                _diag_ns = st.session_state['log_ns']
                if st.button(t('loki_show_pods', ns=_diag_ns)):
                    pod_values = loki.get_label_values('pod', match=f'{{namespace=~"{_diag_ns}"}}')
                    st.write(f"**Hodnoty label `pod` pro namespace `{_diag_ns}`:**", pod_values)
    else:
        st.info(t('loki_info'))

    # --- Namespace selector (cache) ---
    if 'log_ns_list' not in st.session_state:
        with st.spinner(t('spinner_namespaces')):
            namespaces = analyzer.get_namespaces()
            st.session_state['log_ns_list'] = sorted([ns.metadata.name for ns in namespaces])

    ns_names = st.session_state['log_ns_list']
    selected_ns = st.selectbox(t('log_namespace'), ns_names, key="log_ns")

    # --- Pod selector: načíst jen při změně namespace ---
    if st.session_state.get('log_ns_prev') != selected_ns:
        with st.spinner(t('spinner_pods', ns=selected_ns)):
            ns_pods = analyzer.get_pods_in_namespace(selected_ns)
            st.session_state['log_pod_list'] = sorted([p.metadata.name for p in ns_pods])
            st.session_state['log_ns_prev'] = selected_ns
            # Reset downstream selections
            st.session_state.pop('log_pod', None)
            st.session_state.pop('log_container', None)
            st.session_state.pop('log_result', None)

    pod_names = st.session_state.get('log_pod_list', [])
    selected_pod = st.selectbox(t('log_pod'), pod_names if pod_names else [""], key="log_pod")

    # --- Container selector: načíst jen při změně podu ---
    if selected_pod and selected_pod != st.session_state.get('log_pod_prev'):
        containers = analyzer.get_pod_containers(selected_pod, selected_ns)
        st.session_state['log_container_list'] = containers
        st.session_state['log_pod_prev'] = selected_pod
        st.session_state.pop('log_container', None)
        st.session_state.pop('log_result', None)

    containers = st.session_state.get('log_container_list', [])
    if selected_pod and containers:
        selected_container = st.selectbox(t('log_container'), [""] + containers, key="log_container",
                                          help=t('log_container_help'))
    else:
        selected_container = ""

    # --- Parametry ---
    col1, col2, col3 = st.columns(3)
    with col1:
        tail_lines = st.number_input(t('log_lines'), min_value=10, max_value=5000, value=200, step=50, key="log_tail")
    with col2:
        if log_source == 'loki':
            st.number_input(t('log_hours'), min_value=1, max_value=168, value=24, key="log_hours")
    with col3:
        st.write("")  # spacer
        if st.button(t('btn_reset_cache'), help=t('btn_reset_cache_help')):
            for k in ['log_loki_source', 'log_loki_client', 'log_ns_list', 'log_pod_list',
                      'log_ns_prev', 'log_pod_prev', 'log_container_list', 'log_result']:
                st.session_state.pop(k, None)
            st.rerun()

    if selected_pod and st.button(t('btn_load_logs')):
        container_arg = selected_container if selected_container else None
        with st.spinner(t('spinner_logs')):
            logs_text = None
            log_meta = ""

            # Pokud Loki nezná tento namespace → přeskočit rovnou na K8s API
            ns_in_loki = selected_ns in loki_namespaces if loki_namespaces else False
            effective_source = 'loki' if (log_source == 'loki' and ns_in_loki) else 'k8s'
            if log_source == 'loki' and not ns_in_loki:
                st.info(t('loki_ns_not_covered', ns=selected_ns, covered=', '.join(sorted(loki_namespaces)[:8])))

            if effective_source == 'loki':
                lines = loki.get_pod_logs(
                    namespace=selected_ns,
                    pod=selected_pod,
                    container=container_arg,
                    tail_lines=int(tail_lines),
                    hours=int(st.session_state.get('log_hours', 24)),
                    cluster=loki_cluster,
                )
                if lines and lines[0].startswith('[Loki error]'):
                    # Loki chyba → fallback K8s API
                    st.warning(t('loki_fallback_err', e=lines[0]))
                    logs_text = analyzer.get_pod_logs(selected_pod, selected_ns,
                                                      container=container_arg,
                                                      tail_lines=int(tail_lines))
                    log_meta = f"K8s API (fallback) | {selected_ns}/{selected_pod}"
                elif lines:
                    logs_text = '\n'.join(lines)
                    log_meta = f"Loki | {selected_ns}/{selected_pod} | {len(lines)} lines"
                else:
                    # Loki prázdný výsledek → fallback K8s API
                    st.info(t('loki_empty_fallback', h=st.session_state.get('log_hours', 24)))
                    logs_text = analyzer.get_pod_logs(selected_pod, selected_ns,
                                                      container=container_arg,
                                                      tail_lines=int(tail_lines))
                    log_meta = f"K8s API (Loki empty) | {selected_ns}/{selected_pod}"
            else:
                logs_text = analyzer.get_pod_logs(selected_pod, selected_ns,
                                                   container=container_arg,
                                                   tail_lines=int(tail_lines))
                log_meta = f"K8s API | {selected_ns}/{selected_pod}"

            # Normalize: K8s API může vrátit "" → "(žádné logy)"
            if not logs_text:
                logs_text = t('log_no_logs')

            st.session_state['log_result'] = logs_text
            st.session_state['log_result_meta'] = log_meta

    # --- Zobrazit logy (přetrvají i po rerun) ---
    if 'log_result' in st.session_state:
        result = st.session_state['log_result']
        st.caption(st.session_state.get('log_result_meta', ''))
        if result == t('log_no_logs'):
            st.info(t('log_empty_info'))
        else:
            st.code(result, language='log')


with tab7:
    render_logs_tab()


@st.fragment
def render_egress_tab():
    st.header(t('h_egress'))
    st.markdown(t('egress_desc'))

    if not has_cluster_connection():
        st.info(t('not_connected'))
        return

    analyzer = get_analyzer()

    # ── Namespace selector ──────────────────────────────────────────────────
    if 'egress_ns_list' not in st.session_state:
        with st.spinner(t('spinner_namespaces')):
            nss = analyzer.get_namespaces()
            st.session_state['egress_ns_list'] = sorted([ns.metadata.name for ns in nss])

    col_ns, col_reset = st.columns([5, 1])
    with col_ns:
        selected_ns = st.selectbox(t('log_namespace'), st.session_state['egress_ns_list'], key="egress_ns")
    with col_reset:
        st.write("")
        if st.button("🔄", key="egress_ns_reset", help=t('btn_reset_cache_help')):
            st.session_state.pop('egress_ns_list', None)
            st.session_state.pop('egress_pod_list', None)
            st.rerun(scope="fragment")

    # ── Pod IPs in namespace ────────────────────────────────────────────────
    if st.session_state.get('egress_ns_prev') != selected_ns:
        pod_ips = analyzer.get_pod_ips_in_namespace(selected_ns)
        st.session_state['egress_pod_ips'] = pod_ips
        st.session_state['egress_ns_prev'] = selected_ns

    pod_ips = st.session_state.get('egress_pod_ips', [])
    running_pods = [p for p in pod_ips if p['phase'] == 'Running']

    if pod_ips:
        with st.expander(t('egress_pod_ips_title', n=len(pod_ips)), expanded=False):
            st.table([
                {"Pod": p['name'], "IP": p['ip'], "Phase": p['phase'], "Node": p['node']}
                for p in pod_ips
            ])

    # ── CiliumEgressGatewayPolicy lookup ────────────────────────────────────
    with st.spinner(t('egress_loading_policies')):
        try:
            policies = analyzer.get_egress_gateway_policies(namespace=selected_ns)
        except Exception as _e:
            policies = []
            st.warning(t('egress_policy_err', e=str(_e)))

    if policies:
        st.success(t('egress_policies_found', n=len(policies)))
        for pol in policies:
            spec = pol.get("spec", {})
            meta = pol.get("metadata", {})
            egress_ip = spec.get("egressGateway", {}).get("egressIP", "")
            node_sel = spec.get("egressGateway", {}).get("nodeSelector", {})
            interface = spec.get("egressGateway", {}).get("interface", "")
            dest_cidrs = spec.get("destinationCIDRs", [])
            pol_name = meta.get("name", "?")
            with st.expander(f"📋 {pol_name}  →  VIP: **{egress_ip or '?'}**"):
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown(f"**Egress VIP:** `{egress_ip}`")
                    st.markdown(f"**Interface:** `{interface or 'default'}`")
                with col2:
                    st.markdown(f"**Node selector:** `{node_sel}`")
                    if dest_cidrs:
                        st.markdown(f"**Destination CIDRs:** {', '.join(dest_cidrs)}")
                # Verification commands from skill
                with st.expander(t('egress_commands_title'), expanded=False):
                    st.code(
                        f"# 1. Egress VIP a gateway node\n"
                        f"kubectl get ciliumegressgatewaypolicy {pol_name} -o yaml | grep -E 'egressIP|nodeSelector' -A2\n"
                        f"kubectl get nodes -L cilium.io/egress-gateway-node\n\n"
                        f"# 2. Cilium BPF SNAT mapa (na gateway nodu)\n"
                        f"CILIUM_POD=$(kubectl get pod -n kube-system -l app.kubernetes.io/name=cilium-agent \\\n"
                        f"  --field-selector spec.nodeName=<gateway-node> -o jsonpath='{{.items[0].metadata.name}}')\n"
                        f"kubectl exec -n kube-system $CILIUM_POD -- cilium bpf egress list | grep <pod-ip>\n"
                        f"# Správně: 10.35.x.x   0.0.0.0/0   {egress_ip}   <gateway-ip>\n\n"
                        f"# 3. FIB lookup v main tabulce\n"
                        f"kubectl debug node/<gateway-node> -it --image=busybox --profile=sysadmin \\\n"
                        f"  -- chroot /host ip route get <destination-ip>\n"
                        f"# Správně: via <egress-gw> dev ens256\n\n"
                        f"# 4. curl test ze src VIP ({egress_ip})\n"
                        f"kubectl debug node/<gateway-node> -it --image=nicolaka/netshoot --profile=sysadmin \\\n"
                        f"  -- chroot /host curl -v --interface {egress_ip} --max-time 8 https://<dest-ip> -k",
                        language="bash",
                    )
    else:
        st.info(t('egress_no_policies'))

    # ── Connectivity test ───────────────────────────────────────────────────
    st.divider()
    st.subheader(t('egress_test_title'))

    col_target, col_btn = st.columns([4, 1])
    with col_target:
        target_raw = st.text_input(
            t('egress_target_label'),
            placeholder="10.20.30.40:443  nebo  hostname.corp:8080",
            key="egress_target",
            help=t('egress_target_help'),
        )
    with col_btn:
        st.write("")
        run_test = st.button(t('egress_btn_test'), type="primary", key="egress_run_btn")

    if run_test:
        if not target_raw or ':' not in target_raw:
            st.error(t('egress_err_target'))
        else:
            target_parts = target_raw.strip().rsplit(':', 1)
            target_host = target_parts[0]
            target_port = target_parts[1] if len(target_parts) == 2 else "80"

            test_cmd = (
                f"(timeout 5 sh -c 'echo > /dev/tcp/{target_host}/{target_port}' 2>&1 && echo TCP_OPEN)"
                f" || (nc -zv {target_host} {target_port} -w 5 2>&1)"
                f" || echo TCP_CLOSED"
            )

            with st.spinner(f"Spouštím nettest pod v namespace {selected_ns}..."):
                nc_result, pod_node, pod_ip = analyzer.run_pod_command(selected_ns, test_cmd, timeout=40)

            # Interpret result
            success_signals = ["tcp_open", "succeeded", "open", "connected", "200", "ok"]
            fail_signals = ["tcp_closed", "refused", "timed out", "timeout", "no route", "unreachable", "connection reset"]
            output_lower = nc_result.lower()
            is_success = any(s in output_lower for s in success_signals)
            is_fail = any(s in output_lower for s in fail_signals)

            # Parse pod IP from output (fallback if not from K8s status)
            if not pod_ip:
                for _line in nc_result.splitlines():
                    if _line.strip().startswith("POD_IP="):
                        pod_ip = _line.strip().split("=", 1)[1].strip()
                        break

            # Resolve target hostname to IP for Cilium NAT lookup
            import socket as _socket
            try:
                dest_ip_resolved = _socket.gethostbyname(target_host)
            except Exception:
                dest_ip_resolved = target_host

            # Query Cilium BPF egress table — ověř že policy je aktivní v BPF
            # Nový přístup: hledá Egress IP == expected_vip v egress listu (nezávisí na node/pod IP)
            actual_ip = ""
            snat_debug = ""
            expected_vip = ""
            if policies:
                expected_vip = policies[0].get("spec", {}).get("egressGateway", {}).get("egressIP", "")
            with st.spinner("Ověřuji egress policy v Cilium BPF tabulce..."):
                actual_ip, snat_debug = analyzer.get_snat_ip(
                    pod_ip=pod_ip,
                    dest_ip=dest_ip_resolved,
                    dest_port=target_port,
                    node_name=pod_node,
                    expected_vip=expected_vip,
                )

            # Store results in session_state so they persist across reruns
            st.session_state["egress_test_result"] = {
                "nc_result": nc_result,
                "is_success": is_success,
                "is_fail": is_fail,
                "target_host": target_host,
                "target_port": target_port,
                "selected_ns": selected_ns,
                "policies": policies,
                "actual_ip": actual_ip,
                "pod_ip": pod_ip,
                "snat_debug": snat_debug,
            }

    # Render results from session_state (persists when user types in justification input)
    _res = st.session_state.get("egress_test_result")
    if _res:
        _nc_result = _res["nc_result"]
        _is_success = _res["is_success"]
        _is_fail = _res["is_fail"]
        _target_host = _res["target_host"]
        _target_port = _res["target_port"]
        _sel_ns = _res["selected_ns"]
        _policies = _res["policies"]

        if _is_success and not _is_fail:
            st.success(t('egress_result_open', host=_target_host, port=_target_port))
        elif _is_fail:
            st.error(t('egress_result_closed', host=_target_host, port=_target_port))
        else:
            st.warning(t('egress_result_unknown'))

        with st.expander(t('egress_result_detail', pod="nettest-pod")):
            st.code(_nc_result or "(no output)", language="")
            st.caption(f"Spuštěno jako dočasný pod busybox:1.36 v namespace **{_sel_ns}**")
            _snat_debug = _res.get("snat_debug", "")
            if _snat_debug:
                st.markdown("**Cilium SNAT debug:**")
                st.code(_snat_debug, language="")

        # Show egress VIP hint
        _vip = ""
        if _policies:
            vips = [p.get("spec", {}).get("egressGateway", {}).get("egressIP", "")
                    for p in _policies if p.get("spec", {}).get("egressGateway", {}).get("egressIP")]
            if vips:
                _vip = vips[0]
                st.info(t('egress_vip_hint', vips=", ".join(vips)))

        # Show actual outbound IP (from Cilium BPF NAT table)
        _actual_ip = _res.get("actual_ip", "")
        _pod_ip = _res.get("pod_ip", "")
        if _actual_ip:
            if _vip and _actual_ip == _vip:
                st.success(t('egress_actual_ip_match', ip=_actual_ip))
            elif _vip:
                st.warning(t('egress_actual_ip_mismatch', ip=_actual_ip, vip=_vip))
            else:
                st.info(t('egress_actual_ip_info', ip=_actual_ip))
        elif _pod_ip:
            st.caption(t('egress_actual_ip_unknown'))

        # FW ticket generator — only shown when port is closed
        if _is_fail and not _is_success:
            st.divider()
            st.subheader(t('egress_fw_title'))
            st.caption(t('egress_fw_desc'))

            justification = st.text_input(
                t('egress_fw_justification'),
                key="egress_fw_justification_input",
                placeholder="např. Záloha dat do Azure Blob storage",
                help=t('egress_fw_justification_help'),
            )

            if justification:
                # Resolve dest IP if hostname given
                import socket
                dest_ip = _target_host
                try:
                    dest_ip = socket.gethostbyname(_target_host)
                except Exception:
                    pass

                # Pick first policy for VIP and name
                fw_policy_name = _policies[0].get("metadata", {}).get("name", "") if _policies else ""
                fw_vip = ""
                if _policies:
                    fw_vip = _policies[0].get("spec", {}).get("egressGateway", {}).get("egressIP", "")

                xls_bytes = generate_fw_ticket_xls(
                    namespace=_sel_ns,
                    policy_name=fw_policy_name,
                    egress_vip=fw_vip,
                    dest_host=_target_host,
                    dest_ip=dest_ip,
                    port=_target_port,
                    justification=justification,
                )
                filename = f"fw-ticket-{_sel_ns}-{_target_host}-{_target_port}.xlsx"
                st.download_button(
                    label=t('egress_fw_btn'),
                    data=xls_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                )
            else:
                st.warning(t('egress_fw_warn_justification'))


with tab8:
    render_egress_tab()


# Footer
st.markdown("---")
st.markdown(t('footer'))